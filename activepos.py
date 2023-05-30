import os
import time
import requests
from fake_useragent import UserAgent
from openpyxl import Workbook, load_workbook


# User Agent
ua = UserAgent()
url = os.environ.get("URL")
password = os.environ.get('Pass')
headers = {'User-Agent': ua.chrome}

# WorkBook Excel
workbook = Workbook()
sheet = workbook.active


# Connection to main server
resp = requests.get(url, headers=headers, verify=False, timeout=5)
data = resp.json()
sid = data['sid']


# Validations & Get SID
if resp.status_code == 200:
    sid = data['sid']
    print(f"Your sid is {sid}")
else:
    print('No connection to server')


def find_dict(lst, key, value):
    return [d for d in lst if key in d and d[key] == value]


# Transactions
transactions = f"https://10.79.3.10:8080/pos_events?password={password}"
sheet.append(['Reciept No.', 'POS Terminal', 'Price'])
existing_data = []
receipt_counter = 0  # Counter for receipts within an hour
price_total = 0  # Total price within an hour
if os.path.exists('data.xlsx'):
    existing_workbook = load_workbook('data.xlsx', read_only=True)
    existing_sheet = existing_workbook.active

    # Validation of data
    for row in existing_sheet.iter_rows(values_only=True):
        existing_data.append(row)

    existing_workbook.close()

while True:
    transaction = requests.get(transactions, headers=headers, verify=False, timeout=5)
    data = transaction.json()
    transaction.raise_for_status()
    if transaction.status_code == 200:
        print("All is OK!")
    else:
        print("Connection Wrong!")

    target_key = 'type'
    target_value = 'POSNG_RECEIPT_SELL_CLOSE'  # Type is always == "POSNG_RECEIPT_SELL_CLOSE"

    result = find_dict(data, target_key, target_value)

    if result:
        for item in result:
            # Take data
            value1 = item.get('op_id')
            value2 = item.get('pos_terminal_name')
            value3 = item.get('price')

            if value3 is not None:
                value3 = str(value3)
                if len(value3) > 2:
                    value3 = value3[:-2] + '.' + value3[-2:]

            if (value1, value2, value3) not in existing_data:

                # Add data to Excel
                sheet.append([value1, value2, value3])
                existing_data.append((value1, value2, value3))

                receipt_counter += 1
                price_total += float(value3)

        workbook.save('data.xlsx')
        workbook.close()

        if time.localtime().tm_min == 0:  # Every hour

            sheet.append(['Summary', receipt_counter, price_total])

            workbook.save('data.xlsx')
            workbook.close()

            # Reset counters
            receipt_counter = 0
            price_total = 0

        time.sleep(60)
